#!/usr/bin/env python3
"""
Seed the Garden Crawler database from the Excel landscape map.

Reads garden_landscape_map.xlsx and populates:
- orientations (4 rows)
- categories (30 rows from Sheet 1)
- search_phrases (120-180 from category search terms)
- actors (seed actors from Sheet 2)
- people (from Sheet 3)
- campaigns (single "garden-landscape" campaign)

Usage:
    python3 seed_from_excel.py [--excel garden_landscape_map.xlsx] [--db garden.db]
"""
import argparse
import json
import os
import re
import sqlite3
import sys

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_EXCEL = os.path.join(BASE_DIR, "garden_landscape_map.xlsx")
DEFAULT_DB = os.path.join(BASE_DIR, "garden.db")

# The four orientations with full descriptions
ORIENTATIONS = [
    ("GARDEN", "Garden of Eden",
     "Living systems, ecology, land stewardship, regenerative economics, indigenous knowledge, biodiversity, food systems, water, soil. The orientation of tending and cultivation."),
    ("SPACESHIP", "Spaceship Earth",
     "Technology, infrastructure, data, digital twins, AI, blockchain, spatial computing, governance tools, transparency systems, alternative ownership models. The orientation of building and systems design."),
    ("TEMPLE", "Eleusinian Mysteries",
     "Transformative games, LARP, ritual, ceremony, experiential futures, festival culture, consciousness research, performance, art for social change. The orientation of meaning, transformation, and embodied experience."),
    ("ASSEMBLY", "General Assembly of Earth",
     "Global governance innovation, participatory democracy, citizens' assemblies, rights of nature, climate justice, long-termism, planetary health, education for planetary citizenship. The cross-cutting meta-frame where the other three meet."),
]

# Map Excel orientation codes to our codes
ORIENTATION_MAP = {
    "GARDEN": "GARDEN",
    "SPACESHIP": "SPACESHIP",
    "TEMPLE": "TEMPLE",
    "ASSEMBLY": "ASSEMBLY",
}

# Garden-specific intel fields
GARDEN_INTEL_FIELDS = [
    "funding_sources",
    "key_projects",
    "partner_orgs",
    "event_participation",
    "publications",
    "nordic_connection",
    "collaboration_potential",
    "governance_model",
    "open_source_tools",
    "community_size",
]

GARDEN_SCORING_PROMPT = """The Garden is a project to prototype planetary governance through play, simulation, and embodied experience.

We are mapping organisations, companies, networks, research groups, movements, and initiatives worldwide that are — knowingly or not — already doing work that fits within this vision.

The map is organised around four orientations:
- Garden of Eden: Living systems, ecology, land stewardship, regenerative economics
- Spaceship Earth: Technology, infrastructure, governance tools, alternative ownership
- Eleusinian Mysteries: Transformative games, LARP, ritual, experiential futures
- General Assembly of Earth: Global governance innovation, participatory democracy, rights of nature

Rate relevance 1-5:
5 = Directly building governance/simulation/play infrastructure for planetary challenges
4 = Core work in one orientation with clear bridge to others
3 = Strong work in one orientation, potential connector
2 = Adjacent work, useful reference
1 = Tangentially relevant, worth tracking

Include if: actively doing something, work connects to at least one orientation, real outputs (projects, publications, events, products, communities), could plausibly participate in a governance LARP/simulation/assembly.
Exclude if: purely commercial with no governance/ecological/social dimension, defunct (no updates 2+ years), just a blog post not an org, purely AI safety debate without governance application."""


def parse_search_terms(terms_str):
    """Parse semicolon-separated search terms into a list."""
    if not terms_str:
        return []
    return [t.strip() for t in terms_str.split(";") if t.strip()]


def parse_actor_types(types_str):
    """Parse comma-separated actor types."""
    if not types_str:
        return []
    return [t.strip() for t in types_str.split(",") if t.strip()]


def guess_actor_type(type_str):
    """Map free-form type strings from Excel to our enum values."""
    if not type_str:
        return "Network"
    t = type_str.lower().strip()
    if "person" in t:
        return "Person"
    if "company" in t or "tool" in t:
        return "Company"
    if "ngo" in t:
        return "NGO"
    if "research" in t:
        return "Research"
    if "government" in t or "eu" in t or "municipal" in t:
        return "Government"
    if "movement" in t or "declaration" in t or "festival" in t:
        return "Movement"
    if "network" in t or "game" in t or "project" in t:
        return "Network"
    # Multi-type: pick the first recognized one
    for candidate in t.split("/"):
        candidate = candidate.strip()
        if candidate in ("ngo", "company", "research", "government", "network", "movement", "person"):
            return candidate.capitalize()
    return "Network"


def parse_orientation_code(orient_str):
    """Parse orientation string from Excel, handling slashes for multi-orientation."""
    if not orient_str:
        return None, None
    parts = [p.strip().upper() for p in re.split(r'[/,]', orient_str)]
    primary = ORIENTATION_MAP.get(parts[0])
    secondary = ORIENTATION_MAP.get(parts[1]) if len(parts) > 1 else None
    return primary, secondary


def main():
    parser = argparse.ArgumentParser(description="Seed Garden Crawler DB from Excel")
    parser.add_argument("--excel", default=DEFAULT_EXCEL, help="Path to Excel file")
    parser.add_argument("--db", default=DEFAULT_DB, help="Path to SQLite database")
    args = parser.parse_args()

    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
        sys.exit(1)

    # Create DB from schema
    schema_path = os.path.join(BASE_DIR, "schema.sql")
    if not os.path.exists(schema_path):
        print(f"ERROR: schema.sql not found at {schema_path}")
        sys.exit(1)

    conn = sqlite3.connect(args.db)
    conn.execute("PRAGMA foreign_keys = ON")
    cur = conn.cursor()

    with open(schema_path) as f:
        cur.executescript(f.read())
    conn.commit()

    # ─── Orientations ───
    orientation_ids = {}
    for code, name, desc in ORIENTATIONS:
        cur.execute("INSERT OR IGNORE INTO orientations (code, name, description) VALUES (?, ?, ?)",
                    (code, name, desc))
        cur.execute("SELECT id FROM orientations WHERE code = ?", (code,))
        orientation_ids[code] = cur.fetchone()[0]
    print(f"Orientations: {len(orientation_ids)} seeded")

    # ─── Campaign ───
    cur.execute("""INSERT OR IGNORE INTO campaigns (name, description, intel_fields, scoring_prompt)
                   VALUES (?, ?, ?, ?)""",
                ("garden-landscape",
                 "The Garden — Landscape Crawler: mapping actors across four orientations of planetary governance",
                 json.dumps(GARDEN_INTEL_FIELDS),
                 GARDEN_SCORING_PROMPT))
    conn.commit()

    # ─── Read Excel ───
    wb = openpyxl.load_workbook(args.excel)

    # ─── Sheet 1: Categories & Search Terms ───
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    category_ids = {}
    total_phrases = 0

    for row in rows:
        orient_code, cat_name, cat_desc, search_terms_str, actor_types_str = (
            row[0] if len(row) > 0 else None,
            row[1] if len(row) > 1 else None,
            row[2] if len(row) > 2 else None,
            row[3] if len(row) > 3 else None,
            row[4] if len(row) > 4 else None,
        )
        if not cat_name:
            continue

        orient_code = (orient_code or "").strip().upper()
        oid = orientation_ids.get(orient_code)
        if not oid:
            print(f"  WARNING: Unknown orientation '{orient_code}' for category '{cat_name}'")
            continue

        search_terms = parse_search_terms(search_terms_str)
        actor_types = parse_actor_types(actor_types_str)

        cur.execute("""INSERT OR IGNORE INTO categories (orientation_id, name, description, search_terms, actor_types)
                       VALUES (?, ?, ?, ?, ?)""",
                    (oid, cat_name, cat_desc, json.dumps(search_terms), json.dumps(actor_types)))
        cur.execute("SELECT id FROM categories WHERE name = ?", (cat_name,))
        cat_id = cur.fetchone()[0]
        category_ids[cat_name.lower()] = cat_id

        # Insert search phrases
        for term in search_terms:
            cur.execute("INSERT OR IGNORE INTO search_phrases (phrase, category_id, priority) VALUES (?, ?, 'high')",
                        (term, cat_id))
            total_phrases += 1

    print(f"Categories: {len(category_ids)} seeded")
    print(f"Search phrases: {total_phrases} seeded")

    # ─── Sheet 2: Seed Actors ───
    ws = wb[wb.sheetnames[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    actor_count = 0

    for row in rows:
        orient_str, actor_name, type_str, relevance_str, location, source_conn = (
            row[0] if len(row) > 0 else None,
            row[1] if len(row) > 1 else None,
            row[2] if len(row) > 2 else None,
            row[3] if len(row) > 3 else None,
            row[4] if len(row) > 4 else None,
            row[5] if len(row) > 5 else None,
        )
        if not actor_name:
            continue

        primary_code, secondary_code = parse_orientation_code(orient_str)
        primary_oid = orientation_ids.get(primary_code)
        secondary_oid = orientation_ids.get(secondary_code) if secondary_code else None
        actor_type = guess_actor_type(type_str)

        # Parse relevance from description (seed actors get 4-5 by default)
        relevance = 4  # seed actors are pre-selected as relevant

        cur.execute("""INSERT OR IGNORE INTO actors
            (name, type, primary_orientation_id, secondary_orientation_id,
             description, location, scale, maturity, relevance_score,
             connection, contact_pathway, notes)
            VALUES (?, ?, ?, ?, ?, ?, 'Global', 'Active', ?, 'seed', ?, ?)""",
            (actor_name.strip(), actor_type, primary_oid, secondary_oid,
             relevance_str, (location or "").strip(), relevance,
             source_conn, f"Type from Excel: {type_str}"))
        actor_count += 1

    print(f"Seed actors: {actor_count} seeded")

    # ─── Sheet 3: People Map ───
    ws = wb[wb.sheetnames[2]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    people_count = 0

    for row in rows:
        person_name, orient_str, skills, rel_tier, status = (
            row[0] if len(row) > 0 else None,
            row[1] if len(row) > 1 else None,
            row[2] if len(row) > 2 else None,
            row[3] if len(row) > 3 else None,
            row[4] if len(row) > 4 else None,
        )
        if not person_name:
            continue

        # Split name into first/last
        name_parts = person_name.strip().split(None, 1)
        first_name = name_parts[0] if name_parts else ""
        last_name = name_parts[1] if len(name_parts) > 1 else ""

        primary_code, _ = parse_orientation_code(orient_str)
        primary_oid = orientation_ids.get(primary_code)

        # Map relationship tier
        tier = (rel_tier or "").strip()
        valid_tiers = {
            'Core team candidate', 'Strategic advisor', 'Advisor / participant',
            'Participant / contributor', 'Research collaborator', 'To explore',
        }
        # Try to match or default
        matched_tier = None
        for vt in valid_tiers:
            if tier.lower().startswith(vt.lower()[:10]):
                matched_tier = vt
                break
        if not matched_tier:
            if "core" in tier.lower():
                matched_tier = "Core team candidate"
            elif "advisor" in tier.lower() or "strategic" in tier.lower():
                matched_tier = "Strategic advisor"
            elif "research" in tier.lower() or "academic" in tier.lower():
                matched_tier = "Research collaborator"
            elif "participant" in tier.lower():
                matched_tier = "Participant / contributor"
            else:
                matched_tier = "To explore"

        cur.execute("""INSERT OR IGNORE INTO people
            (first_name, last_name, primary_orientation_id, skills, relationship_tier, status, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (first_name, last_name, primary_oid, skills, matched_tier,
             (status or "Not yet contacted").strip(),
             f"Orientation from Excel: {orient_str}"))
        people_count += 1

    print(f"People: {people_count} seeded")

    conn.commit()

    # ─── Summary ───
    print(f"\n{'='*50}")
    print(f"  Garden Crawler database seeded: {args.db}")
    print(f"{'='*50}")
    for table in ['orientations', 'categories', 'actors', 'people', 'search_phrases', 'campaigns']:
        cur.execute(f"SELECT COUNT(*) FROM {table}")
        print(f"  {table}: {cur.fetchone()[0]} rows")

    conn.close()


if __name__ == "__main__":
    main()
